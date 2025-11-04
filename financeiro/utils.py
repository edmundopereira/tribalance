import pandas as pd
from decimal import Decimal
from django.http import HttpResponse
from datetime import datetime
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Lancamento, CategoriaPilar
import locale
# Tenta configurar o locale para pt_BR.UTF-8. Se falhar, tenta pt_BR.
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'pt_BR')
    except locale.Error:
        # Se falhar, usa o padrão do sistema e confia na formatação manual
        pass


def processar_arquivo_importacao(arquivo ):
    """
    Processa arquivo de importação (Excel ou CSV) e retorna resultado.
    RF01 - Importação de Extratos
    """
    resultado = {
        'status': 'SUCESSO',
        'total': 0,
        'importados': 0,
        'duplicados': 0,
        'rejeitados': 0,
        'mensagem': '',
        'erros_detalhados': []  # <-- CORREÇÃO 1: Adiciona lista de erros
    }
    
    try:

        # Força leitura completa da planilha, mesmo que o Excel tenha filtros ou intervalos nomeados
        if arquivo.name.endswith('.csv'):
            df = pd.read_csv(arquivo, encoding='utf-8', sep=None, engine='python')
        elif arquivo.name.endswith(('.xlsx', '.xls')):
            excel = pd.ExcelFile(arquivo)
            df = excel.parse(excel.sheet_names[0], header=0)
        else:
            resultado['status'] = 'ERRO'
            resultado['mensagem'] = 'Formato de arquivo não suportado. Use .csv ou .xlsx'
            return resultado

        
        
        # Validar colunas obrigatórias
        colunas_obrigatorias = ['data', 'lancamento', 'categoria', 'tipo', 'valor', 'fonte', 'conta_final']
        colunas_faltantes = [col for col in colunas_obrigatorias if col not in df.columns]
        
        if colunas_faltantes:
            resultado['status'] = 'ERRO'
            resultado['mensagem'] = f'Colunas obrigatórias faltando: {", ".join(colunas_faltantes)}'
            return resultado
        
        resultado['total'] = len(df)
        
        # RF01.1 - Excluir todos os lançamentos existentes antes de importar
        Lancamento.objects.all().delete()
        
        # Processar cada linha
        for idx, row in df.iterrows():
            try:
                # Converter data (CORREÇÃO 2: Lógica de data mais robusta)
                data = None
                if isinstance(row['data'], str):
                    # Tenta formato YYYY-MM-DD
                    try:
                        data = datetime.strptime(row['data'], '%Y-%m-%d').date()
                    except ValueError:
                        # Tenta formato DD/MM/YYYY
                        try:
                            data = datetime.strptime(row['data'], '%d/%m/%Y').date()
                        except ValueError:
                            # Tenta formato DD-MM-YYYY
                            data = datetime.strptime(row['data'], '%d-%m-%Y').date()
                else:
                    # Usa a conversão padrão do Pandas para objetos datetime
                    data = pd.to_datetime(row['data']).date()
                
                if not data:
                    raise ValueError("Formato de data inválido")

                # Obter mês
                mes_num = data.month
                meses = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
                         'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
                mes = meses[mes_num - 1]
                
                # Validar tipo
                tipo = str(row['tipo']).upper().strip()
                tipos_validos = ['RECEITA', 'DESPESA', 'CRÉDITO', 'DÉBITO']
                if tipo not in tipos_validos:
                    raise ValueError(f"Tipo de lançamento inválido: {tipo}")
                
                # Converter valor
                valor = Decimal(str(row['valor']))
                
                # Classificar pilar
                categoria = str(row['categoria']).strip()
                pilar = classificar_lancamento(categoria)
                
                # A verificação de duplicidade foi removida a pedido do usuário.
                # Todas as transações serão importadas.
                # resultado['duplicados'] += 1
                # continue
                
                # Criar lançamento
                Lancamento.objects.create(
                    data=data,
                    mes=mes,
                    lancamento=str(row['lancamento']).strip(),
                    categoria=categoria,
                    tipo=tipo,
                    valor=valor,
                    fonte=str(row['fonte']).strip(),
                    conta_final=str(row['conta_final']).strip(),
                    pilar_tribalance=pilar
                )
                
                resultado['importados'] += 1
            
            except Exception as e:
                resultado['rejeitados'] += 1
                # CORREÇÃO 3: Captura e registro do erro
                resultado['erros_detalhados'].append(f"Linha {idx + 2}: {str(e)}") 
                continue
        
        # Determinar status final
        if resultado['importados'] > 0 and resultado['rejeitados'] == 0:
            resultado['status'] = 'SUCESSO'
            resultado['mensagem'] = f'Processamento concluído: {resultado["importados"]} importados, {resultado["duplicados"]} duplicados, {resultado["rejeitados"]} rejeitados.'
        elif resultado['importados'] > 0:
            resultado['status'] = 'PARCIAL'
            resultado['mensagem'] = f'Processamento concluído: {resultado["importados"]} importados, {resultado["duplicados"]} duplicados, {resultado["rejeitados"]} rejeitados. (Verifique o log de importação para detalhes dos rejeitados)'
        else:
            resultado['status'] = 'ERRO'
            # CORREÇÃO 4: Inclui os primeiros 5 erros na mensagem principal
            if 'Colunas obrigatórias faltando' in resultado['mensagem']:
                 pass # Se a coluna faltar, a mensagem já foi definida
            else:
                 primeiros_erros = resultado["erros_detalhados"][:5]
                 resultado['mensagem'] = f'Falha na importação. {resultado["rejeitados"]} rejeitados. Primeiros erros: {primeiros_erros}'
        
    except Exception as e:
        resultado['status'] = 'ERRO'
        resultado['mensagem'] = f'Erro fatal ao processar arquivo: {str(e)}'

    # 🧠 Log para diagnóstico (exibe no terminal do runserver)
    print(f"Total lido: {resultado['total']} | Importados: {resultado['importados']} | Rejeitados: {resultado['rejeitados']}")

    return resultado    
    
    
def classificar_lancamento(categoria):
    """
    Classifica um lançamento em um dos pilares do TriBalance.
    RF07 - Classificação de Transações
    """
    categoria = str(categoria).lower().strip()
    
    # Tentar encontrar na base de dados
    try:
        categoria_pilar = CategoriaPilar.objects.get(categoria__iexact=categoria)
        return categoria_pilar.pilar
    except CategoriaPilar.DoesNotExist:
        pass
    
    # Classificação padrão baseada em palavras-chave
    necessidade_palavras = [
        'alimentação', 'aluguel', 'água', 'energia', 'gás', 'internet',
        'telefone', 'transporte', 'combustível', 'saúde', 'medicamento',
        'supermercado', 'mercado', 'padaria', 'farmácia', 'uber', 'táxi',
        'ônibus', 'metrô', 'conta', 'boleto', 'condomínio', 'iptu'
    ]
    
    conforto_palavras = [
        'cinema', 'restaurante', 'barra', 'bar', 'clube', 'academia',
        'viagem', 'hotel', 'passagem', 'diversão', 'lazer', 'jogo',
        'compras', 'roupa', 'sapato', 'beleza', 'cabelo', 'spa',
        'educação', 'curso', 'livro', 'música', 'streaming'
    ]
    
    crescimento_palavras = [
        'investimento', 'aplicação', 'poupança', 'fundo', 'ação', 'bolsa',
        'criptomoeda', 'bitcoin', 'aporte', 'depósito', 'rendimento',
        'juros', 'dividendo', 'resgate', 'empréstimo', 'débito', 'crédito'
    ]
    
    # Verificar palavras-chave
    for palavra in necessidade_palavras:
        if palavra in categoria:
            return 'NECESSIDADE'
    
    for palavra in conforto_palavras:
        if palavra in categoria:
            return 'CONFORTO & EXPERIÊNCIA'
    
    for palavra in crescimento_palavras:
        if palavra in categoria:
            return 'CRESCIMENTO & LIBERDADE'
    
    # Padrão: despesas são necessidade, receitas são crescimento
    return 'NECESSIDADE'


def calcular_kpis(lancamentos):
    """
    Calcula KPIs do dashboard.
    RF02 - Dashboard Financeiro
    """
    from django.db.models import Sum
    
    # Receita total
    # Receita é a soma de todos os valores positivos (valor > 0)
    receita_total = lancamentos.filter(
        valor__gt=0
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    
    # Despesa total
    # Despesa é a soma de todos os valores negativos (valor < 0).
    # Usamos abs() para o KPI, pois o valor deve ser positivo no dashboard.
    despesa_total_query = lancamentos.filter(
        valor__lt=0
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    
    despesa_total = abs(despesa_total_query)
    
    # Saldo
    saldo = receita_total - despesa_total
    
    # Índice de Equilíbrio Financeiro (FBI)
    fbi = float(receita_total / despesa_total) if despesa_total > 0 else 0
    
    # Taxa de Acúmulo Mensal (MWR)
    mwr = float((saldo / receita_total) * 100) if receita_total > 0 else 0
    
    # Taxa de Eficiência de Consumo (CER)
    cer = float((despesa_total / receita_total) * 100) if receita_total > 0 else 0
    
    # Reserva de Segurança (SBI)
    sbi = float(saldo / despesa_total) if despesa_total > 0 else 0
    
    # Progresso de Liberdade Financeira (FPS)
    crescimento = lancamentos.filter(
        pilar_tribalance='CRESCIMENTO & LIBERDADE'
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    fps = float((crescimento / receita_total) * 100) if receita_total > 0 else 0
    
    return {
        'receita_total': float(receita_total),
        'despesa_total': float(despesa_total),
        'saldo': float(saldo),
        'fbi': round(fbi, 2),
        'mwr': round(mwr, 2),
        'cer': round(cer, 2),
        'sbi': round(sbi, 2),
        'fps': round(fps, 2),
    }

def formatar_numero_br(valor, casas_decimais=2):
    """
    Formata um valor numérico para o padrão brasileiro (X.XXX,XX).
    """
    if valor is None:
        return '0,00'
    
    # Converte para float para garantir o funcionamento do locale.format_string
    try:
        valor = float(valor)
    except:
        return '0,00'

    formato = f"%.{casas_decimais}f"
    return locale.format_string(formato, valor, grouping=True).replace('.', '#').replace(',', '.').replace('#', ',')


def formatar_moeda_br(valor):
    """
    Formata um valor numérico para o padrão monetário brasileiro (R$ X.XXX,XX).
    """
    if valor is None:
        return 'R$ 0,00'
    
    # Usa a função formatar_numero_br e adiciona o símbolo de moeda
    return f'R$ {formatar_numero_br(valor)}'


def exportar_csv(lancamentos, nome_arquivo='lancamentos'):
    """
    Exporta lançamentos em formato CSV.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Data', 'Mês', 'Descrição', 'Categoria', 'Tipo',
        'Valor', 'Fonte', 'Conta Final', 'Pilar TriBalance'
    ])
    
    for lancamento in lancamentos:
        writer.writerow([
            lancamento.data.strftime('%d/%m/%Y'),
            lancamento.mes,
            lancamento.lancamento,
            lancamento.categoria,
            lancamento.tipo,
            f'R$ {lancamento.valor:.2f}',
            lancamento.fonte,
            lancamento.conta_final,
            lancamento.pilar_tribalance,
        ])
    
    return response


def exportar_excel(lancamentos, nome_arquivo='lancamentos'):
    """
    Exporta lançamentos em formato Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Lançamentos'
    
    # Cabeçalho
    headers = [
        'Data', 'Mês', 'Descrição', 'Categoria', 'Tipo',
        'Valor', 'Fonte', 'Conta Final', 'Pilar TriBalance'
    ]
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    for row_idx, lancamento in enumerate(lancamentos, 2):
        ws.cell(row=row_idx, column=1).value = lancamento.data.strftime('%d/%m/%Y')
        ws.cell(row=row_idx, column=2).value = lancamento.mes
        ws.cell(row=row_idx, column=3).value = lancamento.lancamento
        ws.cell(row=row_idx, column=4).value = lancamento.categoria
        ws.cell(row=row_idx, column=5).value = lancamento.tipo
        ws.cell(row=row_idx, column=6).value = float(lancamento.valor)
        ws.cell(row=row_idx, column=7).value = lancamento.fonte
        ws.cell(row=row_idx, column=8).value = lancamento.conta_final
        ws.cell(row=row_idx, column=9).value = lancamento.pilar_tribalance
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    # Salvar em memória
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.xlsx"'
    
    wb.save(response)
    return response
